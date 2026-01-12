import openpyxl
import json
from openpyxl import load_workbook

wb = load_workbook(filename="test_json.xlsx")
ws = wb["First Sheet"]

book_dict = {}
book_lst = []
inner_dict = dict()
first_row = ws[1]
temp=[]
tup=tuple()
each_lst = []

for cell in first_row:
    cols = cell.internal_value
    temp.append(cols)
print(temp)

num = ws.max_row-1

j = 0
while j < num:
    i=0
    for row in ws[j+2]:
        # print(row.internal_value)
        tup = (temp[i], row.internal_value)
        print(tup)
        i+=1
        each_lst.append(tup)
    book_lst.append(dict(each_lst))
    # print(book_lst)
    each_lst=[]
    j+=1

print(book_lst)
book_dict["Book"] = book_lst
# print(book_dict)

json_book = json.dumps(book_dict, indent=0)

with open("new.json", "w") as f:
    f.write(json_book)




