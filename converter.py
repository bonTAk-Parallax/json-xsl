import openpyxl
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

json_data = {}

with open("books.json", encoding="utf8") as json_file:
    json_data = json.load(json_file)

wb = Workbook()
ws = wb.active
ws.title = "First Sheet"
lst = []
n = len(json_data["books"][0])-1
lst = list(json_data["books"][0].keys())
print(lst)
# tfnt = 
# for i in range(len(lst)):
#     title = ws.cell(1, i+1, lst[i].capitalize())
#     title.font = Font(name="Arial", size=10, bold=True, color="00FF0000")
#     title.alignment = Alignment(horizontal='center', vertical='center')
#     # ws.merge_cells(start_row=1, start_column=i+1, end_row=1, end_column=i+2)
#     for x in range(n):
#         vals = ws.cell(x+2, i+1, json_data["books"][x][lst[i]])
#         # ws.merge_cells(start_row=x+2, start_column=i+1, end_row=x+3, end_column=i+2)

# ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)


# temp = []
# temp.append(lst[0])
# print(temp)
# j=1
# k=0

# for k in range(len(lst)-1):
#     temp.append(lst[k+1])
#     k+=1
#     print(temp)
#     while j<len(lst)*2:
#         ws.merge_cells(start_row=1, start_column=j, end_row=1, end_column=j+1)
#         title = ws.cell(1, j)
#         try:
#             if len(temp)!=1:
#                 title.value = temp.pop(-2).capitalize()
#             else:
#                 title.value = temp.pop(-1).capitalize()
#         except:
#             break
#         title.font = Font(name="Arial", size=10, bold=True, color="00FF0000")
#         title.alignment = Alignment(horizontal='center', vertical='center')
#         j+=2

def arrange(lst, r, t):
    temp = []
    temp.append(lst[0])
    print(temp)
    j=1
    k=0
    for k in range(len(lst)-1):
        temp.append(lst[k+1])
        k+=1
        print(temp)
        while j<len(lst)*2:
            ws.merge_cells(start_row=r, start_column=j, end_row=r, end_column=j+1)
            title = ws.cell(r, j)
            try:
                if len(temp)!=1:
                    title.value = temp.pop(-2)
                else:
                    title.value = temp.pop(-1)
            except:
                break
            
            if t:
                title.value = title.value.upper()
                title.font = Font(name="Arial", size=10, bold=True, color="00FF0000")
            title.alignment = Alignment(horizontal='center', vertical='center')
            j+=2

arrange(lst, 1, t=1)
vals=[]
# for i in range(len(lst)-1):
#     for x in range(n):
#         vals.append(json_data["books"][x][lst[i]])
#         # vals = [item for item in json_data["books"][x][lst[i]]]
#         # vals = ws.cell(x+2, i+1, json_data["books"][x][lst[i]])
#         arrange(vals,x+1)
#     print(vals)
#     vals.clear


for x in range(n):
    i=0
    while i< len(lst):
        vals.append(json_data["books"][x][lst[i]])
        # vals = [item for item in json_data["books"][x][lst[i]]]
        # vals = ws.cell(x+2, i+1, json_data["books"][x][lst[i]])
        i+=1
    arrange(vals, x+2, 0)
    print(vals)
    vals = []


# ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)


wb.save("test_json.xlsx")